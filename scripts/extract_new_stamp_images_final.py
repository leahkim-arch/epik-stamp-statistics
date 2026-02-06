#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
신규 스탬프 팩 엑셀 파일에서 개별 에셋 이미지를 추출하는 스크립트 (XML 파싱 방식)
"""

import os
import sys
import zipfile
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook
from PIL import Image as PILImage
import io

# 프로젝트 루트 경로
PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_FILE = PROJECT_ROOT / '2026년1월신규스탬프팩.xlsx'
OUTPUT_DIR = PROJECT_ROOT / 'public' / 'new-stamp-assets'

def ensure_output_dir():
    """출력 디렉토리 생성"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"✓ 출력 디렉토리 확인: {OUTPUT_DIR}")

def extract_images_from_xml():
    """XML 파싱을 통해 이미지 추출"""
    if not EXCEL_FILE.exists():
        print(f"❌ 엑셀 파일을 찾을 수 없습니다: {EXCEL_FILE}")
        return False
    
    print(f"\n📂 엑셀 파일 로딩: {EXCEL_FILE.name}")
    
    # 엑셀 파일 구조 읽기
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    worksheet = workbook.active
    
    # 개별 에셋 섹션 찾기
    sections = []
    week_name = None
    
    for row_idx in range(1, worksheet.max_row + 1):
        row = worksheet[row_idx]
        
        # 주차 헤더 찾기
        if len(row) > 1 and row[1].value:
            cell_str = str(row[1].value)
            if '2026년 1월' in cell_str and '주' in cell_str:
                week_name = cell_str.strip()
                continue
        
        # IMG 헤더 찾기
        row_values = [str(cell.value).lower() if cell.value else '' for cell in row]
        if 'img' in row_values and 'title' in row_values:
            img_col_idx = None
            title_col_idx = None
            for col_idx, val in enumerate(row_values, start=1):
                if 'img' in val:
                    img_col_idx = col_idx
                if 'title' in val:
                    title_col_idx = col_idx
            
            if img_col_idx and title_col_idx and week_name:
                sections.append({
                    'header_row': row_idx,
                    'week_name': week_name,
                    'img_col': img_col_idx,
                    'title_col': title_col_idx
                })
    
    print(f"✓ 발견된 개별 에셋 섹션: {len(sections)}개")
    
    # ZIP 파일로 열어서 XML과 이미지 추출
    extracted_count = 0
    skipped_count = 0
    error_count = 0
    
    try:
        with zipfile.ZipFile(EXCEL_FILE, 'r') as zip_ref:
            # drawing XML 읽기
            drawing_xml = zip_ref.read('xl/drawings/drawing1.xml').decode('utf-8')
            root = ET.fromstring(drawing_xml)
            
            # 네임스페이스 정의
            ns = {
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            
            # 이미지-셀 매핑 생성
            image_cell_map = {}  # (row, col) -> (rel_id, image_name)
            
            for anchor in root.findall('.//xdr:oneCellAnchor', ns):
                from_elem = anchor.find('.//xdr:from', ns)
                pic_elem = anchor.find('.//xdr:pic', ns)
                
                if from_elem is not None and pic_elem is not None:
                    col_elem = from_elem.find('xdr:col', ns)
                    row_elem = from_elem.find('xdr:row', ns)
                    
                    if col_elem is not None and row_elem is not None:
                        col_num = int(col_elem.text) + 1  # 0-based to 1-based
                        row_num = int(row_elem.text) + 1  # 0-based to 1-based
                        
                        # 이미지 관계 ID 찾기
                        blip_elem = pic_elem.find('.//a:blip', ns)
                        if blip_elem is not None:
                            rel_id = blip_elem.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                            if rel_id:
                                # 이미지 이름 찾기
                                nv_pic_pr = pic_elem.find('.//xdr:nvPicPr', ns)
                                if nv_pic_pr is not None:
                                    c_nv_pr = nv_pic_pr.find('xdr:cNvPr', ns)
                                    if c_nv_pr is not None:
                                        image_name = c_nv_pr.get('name', '')
                                        image_cell_map[(row_num, col_num)] = (rel_id, image_name)
            
            print(f"✓ XML에서 발견된 이미지 위치: {len(image_cell_map)}개")
            
            # 관계 파일 읽기
            rels_xml = zip_ref.read('xl/drawings/_rels/drawing1.xml.rels').decode('utf-8')
            rels_root = ET.fromstring(rels_xml)
            rel_ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            
            # 관계 ID -> 이미지 파일 경로 매핑
            rel_id_to_image = {}
            for rel in rels_root.findall('.//r:Relationship', rel_ns):
                rel_id = rel.get('Id')
                target = rel.get('Target')
                if 'media' in target:
                    rel_id_to_image[rel_id] = target.replace('../', 'xl/')
            
            print(f"✓ 이미지 관계 매핑: {len(rel_id_to_image)}개")
            
            # 각 섹션 처리
            for section_idx, section in enumerate(sections):
                print(f"\n📸 [{section['week_name']}] 이미지 추출 시작...")
                header_row = section['header_row']
                title_col = section['title_col']
                
                # 다음 섹션까지
                next_section_row = sections[section_idx + 1]['header_row'] if section_idx + 1 < len(sections) else worksheet.max_row + 1
                
                # 주차 번호 추출
                week_num = section['week_name'].replace('2026년 1월 ', '').replace('주', '').strip()
                
                asset_count = 0
                for row_idx in range(header_row + 1, min(header_row + 21, next_section_row)):
                    try:
                        # Title 가져오기
                        title_cell = worksheet.cell(row=row_idx, column=title_col)
                        title = title_cell.value
                        
                        if not title:
                            continue
                        
                        title = str(title).strip()
                        if not title:
                            continue
                        
                        # 파일명 생성
                        safe_title = re.sub(r'[^\w\-_.]', '_', title)
                        filename = f"week{week_num}_{asset_count + 1:02d}_{safe_title}.png"
                        output_path = OUTPUT_DIR / filename
                        
                        # 이미 저장된 파일이면 스킵
                        if output_path.exists():
                            skipped_count += 1
                            asset_count += 1
                            continue
                        
                        # 해당 행, Col 3 (C열)에 이미지가 있는지 확인
                        cell_key = (row_idx, 3)  # Col 3 = C열
                        if cell_key in image_cell_map:
                            rel_id, image_name = image_cell_map[cell_key]
                            
                            if rel_id in rel_id_to_image:
                                image_path = rel_id_to_image[rel_id]
                                
                                try:
                                    # 이미지 파일 읽기
                                    img_data = zip_ref.read(image_path)
                                    
                                    # PIL Image로 변환하여 저장
                                    pil_img = PILImage.open(io.BytesIO(img_data))
                                    pil_img.save(output_path, 'PNG')
                                    
                                    extracted_count += 1
                                    asset_count += 1
                                    print(f"  ✓ [{row_idx}] {filename} (from {image_path})")
                                except Exception as e:
                                    print(f"  ⚠️  [{row_idx}] {filename} - 이미지 저장 실패: {e}")
                                    error_count += 1
                            else:
                                skipped_count += 1
                                asset_count += 1
                        else:
                            skipped_count += 1
                            asset_count += 1
                    
                    except Exception as e:
                        error_count += 1
                        if error_count <= 5:
                            print(f"  ❌ 행 {row_idx} 처리 중 오류: {e}")
                        continue
        
        print(f"\n{'='*60}")
        print(f"📊 추출 결과:")
        print(f"  ✓ 성공: {extracted_count}개")
        print(f"  ⚠️  스킵: {skipped_count}개")
        print(f"  ❌ 오류: {error_count}개")
        print(f"{'='*60}")
        
        return extracted_count > 0
        
    except Exception as e:
        print(f"❌ 파일 처리 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """메인 함수"""
    print("="*60)
    print("🖼️  신규 스탬프 팩 개별 에셋 이미지 추출 스크립트 (Final)")
    print("="*60)
    
    ensure_output_dir()
    
    success = extract_images_from_xml()
    
    if success:
        print("\n✅ 이미지 추출 완료!")
        print(f"📁 저장 위치: {OUTPUT_DIR}")
    else:
        print("\n❌ 이미지 추출 실패")
        sys.exit(1)

if __name__ == '__main__':
    main()
