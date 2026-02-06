#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
신규 스탬프 팩 엑셀 파일에서 개별 에셋 이미지를 추출하는 스크립트 (ZIP 방식)
"""

import os
import sys
import zipfile
import re
from pathlib import Path
from openpyxl import load_workbook
from PIL import Image as PILImage
import io

# 프로젝트 루트 경로
PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_FILE = PROJECT_ROOT / '2026년1월신규스탬프팩.xlsx'
OUTPUT_DIR = PROJECT_ROOT / 'public' / 'new-stamp-assets'

def ensure_output_dir():
    """출력 디렉토리 생성"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"✓ 출력 디렉토리 확인: {OUTPUT_DIR}")

def extract_images_from_zip():
    """엑셀 파일을 ZIP으로 압축 해제하여 이미지 추출"""
    if not EXCEL_FILE.exists():
        print(f"❌ 엑셀 파일을 찾을 수 없습니다: {EXCEL_FILE}")
        return False
    
    print(f"\n📂 엑셀 파일 로딩: {EXCEL_FILE.name}")
    
    # 엑셀 파일 구조 읽기
    workbook = load_workbook(EXCEL_FILE, data_only=True)
    worksheet = workbook.active
    
    # 개별 에셋 섹션 찾기
    sections = []
    week_name = None
    
    for row_idx in range(1, worksheet.max_row + 1):
        row = worksheet[row_idx]
        
        # 주차 헤더 찾기
        if len(row) > 1 and row[1].value:
            cell_str = str(row[1].value)
            if '2026년 1월' in cell_str and '주' in cell_str:
                week_name = cell_str.strip()
                continue
        
        # IMG 헤더 찾기
        row_values = [str(cell.value).lower() if cell.value else '' for cell in row]
        if 'img' in row_values and 'title' in row_values:
            img_col_idx = None
            title_col_idx = None
            for col_idx, val in enumerate(row_values, start=1):
                if 'img' in val:
                    img_col_idx = col_idx
                if 'title' in val:
                    title_col_idx = col_idx
            
            if img_col_idx and title_col_idx and week_name:
                sections.append({
                    'header_row': row_idx,
                    'week_name': week_name,
                    'img_col': img_col_idx,
                    'title_col': title_col_idx
                })
    
    print(f"✓ 발견된 개별 에셋 섹션: {len(sections)}개")
    
    # ZIP 파일로 열어서 이미지 추출
    extracted_count = 0
    skipped_count = 0
    error_count = 0
    
    try:
        with zipfile.ZipFile(EXCEL_FILE, 'r') as zip_ref:
            # xl/media 폴더에서 이미지 파일 찾기
            image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
            print(f"✓ ZIP에서 발견된 이미지 파일: {len(image_files)}개")
            
            # 이미지 파일을 딕셔너리로 저장
            image_data_map = {}
            for img_file in image_files:
                try:
                    img_data = zip_ref.read(img_file)
                    image_data_map[img_file] = img_data
                except Exception as e:
                    print(f"  ⚠️  이미지 파일 읽기 실패 {img_file}: {e}")
            
            # 각 섹션 처리
            for section_idx, section in enumerate(sections):
                print(f"\n📸 [{section['week_name']}] 이미지 추출 시작...")
                header_row = section['header_row']
                title_col = section['title_col']
                
                # 다음 섹션까지
                next_section_row = sections[section_idx + 1]['header_row'] if section_idx + 1 < len(sections) else worksheet.max_row + 1
                
                # 주차 번호 추출
                week_num = section['week_name'].replace('2026년 1월 ', '').replace('주', '').strip()
                
                asset_count = 0
                for row_idx in range(header_row + 1, min(header_row + 21, next_section_row)):
                    try:
                        # Title 가져오기
                        title_cell = worksheet.cell(row=row_idx, column=title_col)
                        title = title_cell.value
                        
                        if not title:
                            continue
                        
                        title = str(title).strip()
                        if not title:
                            continue
                        
                        # 파일명 생성
                        safe_title = re.sub(r'[^\w\-_.]', '_', title)
                        filename = f"week{week_num}_{asset_count + 1:02d}_{safe_title}.png"
                        output_path = OUTPUT_DIR / filename
                        
                        # 이미 저장된 파일이면 스킵
                        if output_path.exists():
                            skipped_count += 1
                            asset_count += 1
                            continue
                        
                        # 해당 행에 연결된 이미지 찾기 (openpyxl의 이미지 객체 사용)
                        # 이미지 객체에서 해당 행의 이미지 찾기
                        found_image = None
                        for img in worksheet._images:
                            if hasattr(img, 'anchor'):
                                anchor = img.anchor
                                if hasattr(anchor, '_from'):
                                    img_row = anchor._from.row + 1
                                    img_col_num = anchor._from.col + 1
                                    if img_row == row_idx and img_col_num == 3:  # C열
                                        found_image = img
                                        break
                        
                        if found_image:
                            try:
                                # 이미지 데이터 추출
                                img_data = found_image._data()
                                
                                # PIL Image로 변환하여 저장
                                pil_img = PILImage.open(io.BytesIO(img_data))
                                pil_img.save(output_path, 'PNG')
                                
                                extracted_count += 1
                                asset_count += 1
                                print(f"  ✓ [{row_idx}] {filename}")
                            except Exception as e:
                                print(f"  ⚠️  [{row_idx}] {filename} - 이미지 저장 실패: {e}")
                                error_count += 1
                        else:
                            # 이미지가 없으면 fallback 이미지 생성하지 않고 스킵
                            skipped_count += 1
                            asset_count += 1
                    
                    except Exception as e:
                        error_count += 1
                        if error_count <= 5:
                            print(f"  ❌ 행 {row_idx} 처리 중 오류: {e}")
                        continue
        
        print(f"\n{'='*60}")
        print(f"📊 추출 결과:")
        print(f"  ✓ 성공: {extracted_count}개")
        print(f"  ⚠️  스킵: {skipped_count}개")
        print(f"  ❌ 오류: {error_count}개")
        print(f"{'='*60}")
        
        return extracted_count > 0
        
    except Exception as e:
        print(f"❌ ZIP 파일 처리 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """메인 함수"""
    print("="*60)
    print("🖼️  신규 스탬프 팩 개별 에셋 이미지 추출 스크립트 (v2)")
    print("="*60)
    
    ensure_output_dir()
    
    success = extract_images_from_zip()
    
    if success:
        print("\n✅ 이미지 추출 완료!")
        print(f"📁 저장 위치: {OUTPUT_DIR}")
    else:
        print("\n❌ 이미지 추출 실패")
        sys.exit(1)

if __name__ == '__main__':
    main()
