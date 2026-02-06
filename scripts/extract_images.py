#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀 파일에서 이미지를 추출하여 thumbnails 폴더에 저장하는 스크립트
"""

import os
import sys
import zipfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io

# 프로젝트 루트 경로
PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_FILE = PROJECT_ROOT / '전체스탬프리스트.xlsx'
OUTPUT_DIR = PROJECT_ROOT / 'public' / 'thumbnails'

def ensure_output_dir():
    """출력 디렉토리 생성"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"✓ 출력 디렉토리 확인: {OUTPUT_DIR}")

def find_columns(worksheet):
    """CategoryName과 Thumbnail 컬럼 인덱스 찾기"""
    header_row = None
    thumbnail_col_idx = None
    category_name_col_idx = None
    
    # 헤더 행 찾기 (최대 10행까지 확인)
    for row_idx in range(1, min(11, worksheet.max_row + 1)):
        row = worksheet[row_idx]
        for col_idx, cell in enumerate(row, start=1):
            if cell.value:
                cell_value = str(cell.value).lower()
                if 'thumbnail' in cell_value or '썸네일' in cell_value:
                    thumbnail_col_idx = col_idx
                if 'categoryname' in cell_value or 'category' in cell_value:
                    category_name_col_idx = col_idx
                    header_row = row_idx
    
    if header_row:
        print(f"✓ 헤더 행: {header_row}")
        print(f"✓ Thumbnail 컬럼: 열 {thumbnail_col_idx}")
        print(f"✓ CategoryName 컬럼: 열 {category_name_col_idx}")
        return header_row, thumbnail_col_idx, category_name_col_idx
    
    return None, None, None

def extract_images_from_excel():
    """엑셀 파일에서 이미지 추출"""
    if not EXCEL_FILE.exists():
        print(f"❌ 엑셀 파일을 찾을 수 없습니다: {EXCEL_FILE}")
        return False
    
    print(f"\n📂 엑셀 파일 로딩: {EXCEL_FILE.name}")
    
    try:
        workbook = load_workbook(EXCEL_FILE, data_only=True)
        worksheet = workbook.active
        
        print(f"✓ 시트명: {worksheet.title}")
        print(f"✓ 총 행 수: {worksheet.max_row}")
        print(f"✓ 총 열 수: {worksheet.max_column}")
        
        # 컬럼 찾기
        header_row, thumbnail_col_idx, category_name_col_idx = find_columns(worksheet)
        
        if not category_name_col_idx:
            print("❌ CategoryName 컬럼을 찾을 수 없습니다.")
            return False
        
        # 이미지 추출
        extracted_count = 0
        skipped_count = 0
        error_count = 0
        
        print(f"\n📸 이미지 추출 시작...")
        
        # 이미지 파일명과 셀 위치 매핑
        image_map = {}
        
        # worksheet._images에서 이미지 정보 추출
        if hasattr(worksheet, '_images') and worksheet._images:
            print(f"✓ 발견된 이미지 객체 수: {len(worksheet._images)}")
            
            for idx, img in enumerate(worksheet._images):
                try:
                    # 이미지의 anchor에서 위치 정보 추출
                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        if hasattr(anchor, '_from'):
                            row = anchor._from.row + 1  # 0-based to 1-based
                            col = anchor._from.col + 1
                            
                            # 이미지 데이터 추출
                            if hasattr(img, '_data'):
                                image_map[row] = {
                                    'image': img,
                                    'row': row,
                                    'col': col
                                }
                except Exception as e:
                    print(f"  ⚠️  이미지 {idx+1} 매핑 오류: {e}")
                    continue
        
        print(f"✓ 매핑된 이미지 객체 수: {len(image_map)}")
        
        # 데이터 행 처리 (헤더 다음 행부터)
        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            try:
                # CategoryName 컬럼에서 파일명 가져오기
                category_name_cell = worksheet.cell(row=row_idx, column=category_name_col_idx)
                category_name = category_name_cell.value
                
                if not category_name:
                    skipped_count += 1
                    continue
                
                # 파일명 정리
                category_name = str(category_name).strip()
                if not category_name:
                    skipped_count += 1
                    continue
                
                # 파일명 생성 (CategoryName + .png)
                safe_filename = f"{category_name}.png"
                # 파일명에서 특수문자 제거 (파일 시스템 호환)
                safe_filename = "".join(c for c in safe_filename if c.isalnum() or c in ('_', '-', '.'))
                
                output_path = OUTPUT_DIR / safe_filename
                
                # 이미 이미 저장된 파일이면 스킵
                if output_path.exists():
                    skipped_count += 1
                    continue
                
                # 해당 행에 이미지가 있는지 확인
                if row_idx in image_map:
                    img_info = image_map[row_idx]
                    img = img_info['image']
                    try:
                        # 이미지 데이터 추출
                        img_data = img._data()
                        
                        # PIL Image로 변환하여 저장
                        pil_img = PILImage.open(io.BytesIO(img_data))
                        
                        # PNG로 저장
                        pil_img.save(output_path, 'PNG')
                        
                        extracted_count += 1
                        if extracted_count <= 10 or extracted_count % 50 == 0:
                            print(f"  ✓ [{row_idx}] {safe_filename}")
                    except Exception as e:
                        print(f"  ⚠️  [{row_idx}] {safe_filename} - 이미지 저장 실패: {e}")
                        error_count += 1
                else:
                    skipped_count += 1
                    # 이미지가 없는 행은 스킵
            
            except Exception as e:
                error_count += 1
                if error_count <= 5:
                    print(f"  ❌ 행 {row_idx} 처리 중 오류: {e}")
                continue
        
        print(f"\n{'='*60}")
        print(f"📊 추출 결과:")
        print(f"  ✓ 성공: {extracted_count}개")
        print(f"  ⚠️  스킵: {skipped_count}개")
        print(f"  ❌ 오류: {error_count}개")
        print(f"{'='*60}")
        
        return extracted_count > 0
        
    except Exception as e:
        print(f"❌ 엑셀 파일 처리 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """메인 함수"""
    print("="*60)
    print("🖼️  엑셀 이미지 추출 스크립트")
    print("="*60)
    
    ensure_output_dir()
    
    success = extract_images_from_excel()
    
    if success:
        print("\n✅ 이미지 추출 완료!")
        print(f"📁 저장 위치: {OUTPUT_DIR}")
    else:
        print("\n❌ 이미지 추출 실패")
        sys.exit(1)

if __name__ == '__main__':
    main()
