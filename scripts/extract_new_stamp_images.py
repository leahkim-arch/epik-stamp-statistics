#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
신규 스탬프 팩 엑셀 파일에서 개별 에셋 이미지를 추출하는 스크립트
"""

import os
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import re

# 프로젝트 루트 경로
PROJECT_ROOT = Path(__file__).parent.parent
EXCEL_FILE = PROJECT_ROOT / '2026년1월신규스탬프팩.xlsx'
OUTPUT_DIR = PROJECT_ROOT / 'public' / 'new-stamp-assets'

def ensure_output_dir():
    """출력 디렉토리 생성"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"✓ 출력 디렉토리 확인: {OUTPUT_DIR}")

def find_asset_sections(worksheet):
    """개별 에셋 섹션의 헤더 행 찾기"""
    sections = []
    week_name = None
    
    for row_idx in range(1, worksheet.max_row + 1):
        row = worksheet[row_idx]
        
        # 주차 헤더 찾기 (B열에 "2026년 1월 X주")
        if len(row) > 1 and row[1].value:
            cell_str = str(row[1].value)
            if '2026년 1월' in cell_str and '주' in cell_str:
                week_name = cell_str.strip()
                continue
        
        # "IMG"와 "Title"이 모두 있는 행 찾기
        row_values = [str(cell.value).lower() if cell.value else '' for cell in row]
        if 'img' in row_values and 'title' in row_values:
            img_col_idx = None
            title_col_idx = None
            for col_idx, val in enumerate(row_values, start=1):
                if 'img' in val:
                    img_col_idx = col_idx
                if 'title' in val:
                    title_col_idx = col_idx
            
            if img_col_idx and title_col_idx and week_name:
                sections.append({
                    'header_row': row_idx,
                    'week_name': week_name,
                    'img_col': img_col_idx,
                    'title_col': title_col_idx
                })
    
    return sections

def extract_images_from_excel():
    """엑셀 파일에서 개별 에셋 이미지 추출"""
    if not EXCEL_FILE.exists():
        print(f"❌ 엑셀 파일을 찾을 수 없습니다: {EXCEL_FILE}")
        return False
    
    print(f"\n📂 엑셀 파일 로딩: {EXCEL_FILE.name}")
    
    try:
        workbook = load_workbook(EXCEL_FILE, data_only=True)
        worksheet = workbook.active
        
        print(f"✓ 시트명: {worksheet.title}")
        print(f"✓ 총 행 수: {worksheet.max_row}")
        
        # 이미지 매핑 (행 번호 -> 이미지 객체)
        image_map = {}
        
        if hasattr(worksheet, '_images') and worksheet._images:
            print(f"✓ 발견된 이미지 객체 수: {len(worksheet._images)}")
            
            for idx, img in enumerate(worksheet._images):
                try:
                    if hasattr(img, 'anchor'):
                        anchor = img.anchor
                        if hasattr(anchor, '_from'):
                            row = anchor._from.row + 1  # 0-based to 1-based
                            col = anchor._from.col + 1
                            
                            if row not in image_map:
                                image_map[row] = []
                            
                            image_map[row].append({
                                'image': img,
                                'row': row,
                                'col': col
                            })
                except Exception as e:
                    print(f"  ⚠️  이미지 {idx+1} 매핑 오류: {e}")
                    continue
        
        print(f"✓ 매핑된 이미지 행 수: {len(image_map)}")
        
        # 개별 에셋 섹션 찾기
        sections = find_asset_sections(worksheet)
        print(f"\n✓ 발견된 개별 에셋 섹션: {len(sections)}개")
        
        extracted_count = 0
        skipped_count = 0
        error_count = 0
        
        # 각 섹션 처리
        for section_idx, section in enumerate(sections):
            print(f"\n📸 [{section['week_name']}] 이미지 추출 시작...")
            header_row = section['header_row']
            img_col = section['img_col']
            title_col = section['title_col']
            
            # 다음 섹션까지 또는 파일 끝까지
            next_section_row = sections[section_idx + 1]['header_row'] if section_idx + 1 < len(sections) else worksheet.max_row + 1
            
            # 최대 20개 에셋만 처리
            asset_count = 0
            for row_idx in range(header_row + 1, min(header_row + 21, next_section_row)):
                try:
                    # Title 컬럼에서 파일명 가져오기
                    title_cell = worksheet.cell(row=row_idx, column=title_col)
                    title = title_cell.value
                    
                    if not title:
                        continue
                    
                    title = str(title).strip()
                    if not title:
                        continue
                    
                    # 파일명 생성 (Title 기반)
                    safe_title = re.sub(r'[^\w\-_.]', '_', title)
                    # 주차명을 간단하게 변환 (예: "2026년 1월 1주" -> "week1")
                    week_num = section['week_name'].replace('2026년 1월 ', '').replace('주', '').strip()
                    filename = f"week{week_num}_{asset_count + 1:02d}_{safe_title}.png"
                    
                    output_path = OUTPUT_DIR / filename
                    
                    # 이미 저장된 파일이면 스킵
                    if output_path.exists():
                        skipped_count += 1
                        asset_count += 1
                        continue
                    
                    # 해당 행에 이미지가 있는지 확인 (Col 3 = C열)
                    if row_idx in image_map:
                        img_infos = image_map[row_idx]
                        # IMG 컬럼(Col 3)에 가장 가까운 이미지 찾기
                        img_info = None
                        min_col_diff = float('inf')
                        
                        for info in img_infos:
                            col_diff = abs(info['col'] - img_col)
                            if col_diff < min_col_diff:
                                min_col_diff = col_diff
                                img_info = info
                        
                        if img_info and min_col_diff <= 2:  # Col 3 근처 (1-2 열 차이)
                            try:
                                img = img_info['image']
                                # 이미지 데이터 추출
                                img_data = img._data()
                                
                                # PIL Image로 변환하여 저장
                                pil_img = PILImage.open(io.BytesIO(img_data))
                                
                                # PNG로 저장
                                pil_img.save(output_path, 'PNG')
                                
                                extracted_count += 1
                                asset_count += 1
                                print(f"  ✓ [{row_idx}] {filename}")
                            except Exception as e:
                                print(f"  ⚠️  [{row_idx}] {filename} - 이미지 저장 실패: {e}")
                                error_count += 1
                        else:
                            skipped_count += 1
                    else:
                        skipped_count += 1
                    
                    asset_count += 1
                
                except Exception as e:
                    error_count += 1
                    if error_count <= 5:
                        print(f"  ❌ 행 {row_idx} 처리 중 오류: {e}")
                    continue
        
        print(f"\n{'='*60}")
        print(f"📊 추출 결과:")
        print(f"  ✓ 성공: {extracted_count}개")
        print(f"  ⚠️  스킵: {skipped_count}개")
        print(f"  ❌ 오류: {error_count}개")
        print(f"{'='*60}")
        
        return extracted_count > 0
        
    except Exception as e:
        print(f"❌ 엑셀 파일 처리 중 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """메인 함수"""
    print("="*60)
    print("🖼️  신규 스탬프 팩 개별 에셋 이미지 추출 스크립트")
    print("="*60)
    
    ensure_output_dir()
    
    success = extract_images_from_excel()
    
    if success:
        print("\n✅ 이미지 추출 완료!")
        print(f"📁 저장 위치: {OUTPUT_DIR}")
    else:
        print("\n❌ 이미지 추출 실패")
        sys.exit(1)

if __name__ == '__main__':
    main()
